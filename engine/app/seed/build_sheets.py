"""Generate the procurement spreadsheet from the seed JSON.

PRD 2.2.3 says the PO side is "seeded from CSV to honour the 'spreadsheet'
framing" — the brief's premise is that procurement keeps its POs in a
spreadsheet and AP staff look them up by hand. So the CSVs here are the real
seed source the engine loads, not a decorative export.

Also writes an .xlsx workbook of the same data for anyone who would rather open
it in Excel than read a CSV.

Run with:  python -m app.seed.build_sheets
"""

from __future__ import annotations

import csv
import json
from pathlib import Path

HERE = Path(__file__).resolve().parent

VENDOR_COLUMNS = [
    "id", "vendor_code", "legal_name", "trade_name", "tax_id", "status",
    "approval_status", "default_currency", "permitted_currencies",
    "payment_terms_days", "registered_address", "aliases",
    "contract_start", "contract_end", "bank_account_hash", "_seed_note",
]

PO_COLUMNS = [
    "id", "po_number", "vendor_id", "vendor_name", "status", "currency",
    "subtotal", "tax_amount", "total_amount", "po_date", "valid_from",
    "valid_until", "allows_partial_invoicing", "amount_tolerance_pct",
    "amount_tolerance_abs", "cost_center", "approved_by", "approved_at",
    "_seed_note",
]

PO_LINE_COLUMNS = [
    "id", "po_id", "po_number", "line_no", "sku", "description",
    "quantity_ordered", "uom", "unit_price", "line_total", "tax_rate_pct",
    "hsn_sac_code",
]

# Lists are stored pipe-separated so a single CSV cell survives a round trip
# through Excel without needing quoting rules nobody remembers.
LIST_FIELDS = {"aliases", "permitted_currencies"}


def _flatten(value):
    if isinstance(value, list):
        return " | ".join(str(v) for v in value)
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if value is None:
        return ""
    return str(value)


def write_csv(path: Path, columns: list[str], rows: list[dict]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({c: _flatten(row.get(c)) for c in columns})


def main() -> int:
    vendors = json.loads((HERE / "vendors.json").read_text())
    pos = json.loads((HERE / "purchase_orders.json").read_text())
    po_lines = json.loads((HERE / "po_lines.json").read_text())

    vendor_names = {v["id"]: v["trade_name"] for v in vendors}
    po_numbers = {p["id"]: p["po_number"] for p in pos}

    # Denormalise the human-readable name into the sheet. A spreadsheet a person
    # actually reads should not make them join two tabs by ID.
    pos_out = [{**p, "vendor_name": vendor_names.get(p["vendor_id"], "")} for p in pos]
    lines_out = [{**l, "po_number": po_numbers.get(l["po_id"], "")} for l in po_lines]

    write_csv(HERE / "vendors.csv", VENDOR_COLUMNS, vendors)
    write_csv(HERE / "purchase_orders.csv", PO_COLUMNS, pos_out)
    write_csv(HERE / "po_lines.csv", PO_LINE_COLUMNS, lines_out)
    print(f"  vendors.csv          {len(vendors):3} rows")
    print(f"  purchase_orders.csv  {len(pos):3} rows")
    print(f"  po_lines.csv         {len(po_lines):3} rows")

    try:
        _write_workbook(vendors, pos_out, lines_out)
    except ImportError:
        print("\n  (openpyxl not installed — skipped the .xlsx workbook;")
        print("   the CSVs open in Excel directly.)")
    return 0


def _write_workbook(vendors, pos, po_lines) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    sheets = [
        ("Purchase Orders", PO_COLUMNS, pos),
        ("PO Lines", PO_LINE_COLUMNS, po_lines),
        ("Vendors", VENDOR_COLUMNS, vendors),
    ]

    for title, columns, rows in sheets:
        ws = wb.create_sheet(title)
        ws.append([c.replace("_", " ").strip().title() for c in columns])
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"

        for row in rows:
            ws.append([_flatten(row.get(c)) for c in columns])

        for index, column in enumerate(columns, start=1):
            width = max(
                [len(column)] + [len(_flatten(r.get(column))) for r in rows]
            )
            ws.column_dimensions[get_column_letter(index)].width = min(max(width + 2, 10), 46)

        ws.auto_filter.ref = ws.dimensions

    readme = wb.create_sheet("Read me", 0)
    for line in [
        ["Procurement master data"],
        [],
        ["This workbook is the procurement system the invoice platform looks POs up in."],
        ["It stands in for the spreadsheet the brief describes, and it is the actual"],
        ["seed source the engine loads at startup — not an export."],
        [],
        ["Purchase Orders", "One row per PO. total_amount is what PO-07 measures cumulative billing against."],
        ["", "allows_partial_invoicing drives POL-02 and is what makes Edge Case 1 legitimate."],
        ["PO Lines", "One row per ordered line. unit_price is the contracted price LIN-03 checks against."],
        ["", "quantity_ordered is the ceiling LIN-02 measures cumulative billed quantity against."],
        ["Vendors", "Vendor master. VEN-01 resolves an invoice to a row here by tax ID, name, alias or fuzzy match."],
        [],
        ["Invoices are NEVER stored here. They arrive as PDFs and are matched against this data."],
        ["One PO can carry many invoices; the running balance lives in the platform's"],
        ["consumption ledger, not in this sheet."],
    ]:
        readme.append(line)
    readme["A1"].font = Font(bold=True, size=13)
    readme.column_dimensions["A"].width = 20
    readme.column_dimensions["B"].width = 96

    path = HERE / "procurement_master.xlsx"
    wb.save(path)
    print(f"  procurement_master.xlsx  ({len(sheets)} sheets + read me)")


if __name__ == "__main__":
    raise SystemExit(main())
